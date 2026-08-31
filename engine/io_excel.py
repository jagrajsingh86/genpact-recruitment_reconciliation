"""Excel loading and strict header validation for the reconciliation engine."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from os import PathLike
from typing import BinaryIO, Iterable, Literal

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .reconcile import (
    FINDING_COLUMNS,
    FieldPair,
    ReconciliationConfig,
    ReconciliationResult,
    ValueNormaliser,
    reconcile,
)


WorkbookSource = str | PathLike[str] | BinaryIO
WorkbookDestination = str | PathLike[str] | BinaryIO

FIXED_REQUISITION_COLUMNS = (
    "Position Number",
    "Requisition No",
    "Recruiter (R) Name",
    "Recruitment Stage",
    "Current Status",
)
FIXED_POSITION_COLUMNS = ("Position Number",)
FIELD_MAP_COLUMNS = ("req_col", "pos_col", "rule", "weight", "status", "notes")
NORMALISER_COLUMNS = ("report", "column", "pattern", "replacement")
RESOLVED_COLUMNS = ("Requisition No", "Field", "Exception Type", "Status")

REGISTER_SHEET = "Mismatch_Register"
SUMMARY_SHEET = "Run_Summary"
RESOLVED_SHEET = "Resolved_This_Run"

PROVENANCE_TEXT = {
    "library": "SYNTHETIC - demo only. Never imply access to ANZ data.",
    "upload": "UPLOADED - data-handling gate applies",
}

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="282A27")
_HEADER_FONT = Font(name="Aptos", bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Aptos", size=10)
_TEXT_NUMBER_FORMAT = "@"


class HeaderValidationError(ValueError):
    """Raised when an extract omits one or more contract-required headers."""


@dataclass(frozen=True)
class LoadedInputs:
    requisitions: tuple[dict[str, str], ...]
    positions: tuple[dict[str, str], ...]
    config: ReconciliationConfig


def _read_sheet(source: WorkbookSource, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(
        source,
        sheet_name=sheet_name,
        dtype=str,
        keep_default_na=False,
        engine="openpyxl",
    )


def _missing(actual: Iterable[object], required: Iterable[str]) -> list[str]:
    actual_set = set(actual)
    return [column for column in required if column not in actual_set]


def _require_headers(frame: pd.DataFrame, required: Iterable[str], label: str) -> None:
    missing = _missing(frame.columns, required)
    if missing:
        raise HeaderValidationError(f"{label} missing required columns: {', '.join(missing)}")


def _first_data_sheet(source: WorkbookSource) -> str:
    with pd.ExcelFile(source, engine="openpyxl") as workbook:
        data_sheets = [name for name in workbook.sheet_names if name != "_SYNTHETIC"]
    if not data_sheets:
        raise ValueError("Extract workbook has no data sheet (only _SYNTHETIC was found)")
    return data_sheets[0]


def read_extract(source: WorkbookSource) -> pd.DataFrame:
    """Read the first non-provenance sheet with every cell represented as text."""

    return _read_sheet(source, _first_data_sheet(source))


def read_config(source: WorkbookSource) -> ReconciliationConfig:
    with pd.ExcelFile(source, engine="openpyxl") as workbook:
        missing_sheets = [
            sheet for sheet in ("FieldMap", "ValueNormalisers") if sheet not in workbook.sheet_names
        ]
        if missing_sheets:
            raise ValueError(f"Config workbook missing required sheets: {', '.join(missing_sheets)}")
        field_map = pd.read_excel(
            workbook,
            sheet_name="FieldMap",
            dtype=str,
            keep_default_na=False,
        )
        normalisers = pd.read_excel(
            workbook,
            sheet_name="ValueNormalisers",
            dtype=str,
            keep_default_na=False,
        )
    _require_headers(field_map, FIELD_MAP_COLUMNS, "Config FieldMap")
    _require_headers(normalisers, NORMALISER_COLUMNS, "Config ValueNormalisers")
    return ReconciliationConfig(
        field_map=tuple(FieldPair(**row) for row in field_map.loc[:, FIELD_MAP_COLUMNS].to_dict("records")),
        normalisers=tuple(
            ValueNormaliser(**row)
            for row in normalisers.loc[:, NORMALISER_COLUMNS].to_dict("records")
        ),
    )


def validate_headers(
    requisitions: pd.DataFrame,
    positions: pd.DataFrame,
    config: ReconciliationConfig,
) -> None:
    confirmed = config.confirmed_pairs
    requisition_required = (*FIXED_REQUISITION_COLUMNS, *(pair.req_col for pair in confirmed))
    position_required = (*FIXED_POSITION_COLUMNS, *(pair.pos_col for pair in confirmed))
    _require_headers(requisitions, dict.fromkeys(requisition_required), "Requisition report")
    _require_headers(positions, dict.fromkeys(position_required), "Position report")


def load_inputs(
    requisition_source: WorkbookSource,
    position_source: WorkbookSource,
    config_source: WorkbookSource,
) -> LoadedInputs:
    config = read_config(config_source)
    requisitions = read_extract(requisition_source)
    positions = read_extract(position_source)
    validate_headers(requisitions, positions, config)
    return LoadedInputs(
        requisitions=tuple(requisitions.to_dict("records")),
        positions=tuple(positions.to_dict("records")),
        config=config,
    )


def read_previous_findings(source: WorkbookSource) -> tuple[dict[str, str], ...]:
    """Load current findings from a previously produced register workbook."""

    with pd.ExcelFile(source, engine="openpyxl") as workbook:
        if REGISTER_SHEET not in workbook.sheet_names:
            raise ValueError(f"Previous register missing required sheet: {REGISTER_SHEET}")
        frame = pd.read_excel(
            workbook,
            sheet_name=REGISTER_SHEET,
            dtype=str,
            keep_default_na=False,
        )
    _require_headers(frame, FINDING_COLUMNS, f"Previous register {REGISTER_SHEET}")
    return tuple(frame.loc[:, FINDING_COLUMNS].to_dict("records"))


def reconcile_workbooks(
    requisition_source: WorkbookSource,
    position_source: WorkbookSource,
    config_source: WorkbookSource,
    run_date: str,
    previous_register_source: WorkbookSource | None = None,
    *,
    apply_normalisers: bool = True,
) -> ReconciliationResult:
    """Run reconciliation directly from upload-compatible workbook sources."""

    loaded = load_inputs(requisition_source, position_source, config_source)
    previous = (
        read_previous_findings(previous_register_source)
        if previous_register_source is not None
        else ()
    )
    return reconcile(
        loaded.requisitions,
        loaded.positions,
        loaded.config,
        run_date,
        previous,
        apply_normalisers=apply_normalisers,
    )


def _append_rows(
    sheet: Worksheet,
    columns: Iterable[str],
    rows: Iterable[dict[str, object]],
) -> None:
    ordered_columns = tuple(columns)
    sheet.append(ordered_columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in ordered_columns])


def _style_tabular_sheet(
    sheet: Worksheet,
    *,
    text_columns: Iterable[str] = (),
) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    if sheet.max_column:
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 22

    header_by_name = {cell.value: cell.column for cell in sheet[1]}
    for column_name in text_columns:
        column_index = header_by_name.get(column_name)
        if column_index is None:
            continue
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.value = "" if cell.value is None else str(cell.value)
            cell.number_format = _TEXT_NUMBER_FORMAT

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = _BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    for column_index in range(1, sheet.max_column + 1):
        values = [sheet.cell(row=row, column=column_index).value for row in range(1, sheet.max_row + 1)]
        width = min(max((len(str(value or "")) for value in values), default=0) + 2, 55)
        sheet.column_dimensions[get_column_letter(column_index)].width = max(width, 12)


def _summary_rows(result: ReconciliationResult, data_mode: str) -> tuple[dict[str, object], ...]:
    if data_mode not in PROVENANCE_TEXT:
        allowed = ", ".join(sorted(PROVENANCE_TEXT))
        raise ValueError(f"Unknown data mode {data_mode!r}; expected one of: {allowed}")
    summary = result.summary
    return (
        {"Metric": "RUN DATE", "Value": summary["run_date"]},
        {"Metric": "ACTIVE REQUISITIONS", "Value": summary["active_count"]},
        {"Metric": "FIELD PAIRS CHECKED", "Value": summary["pairs_checked"]},
        {"Metric": "EXCEPTIONS", "Value": summary["exceptions"]},
        {"Metric": "HIGH", "Value": summary["high"]},
        {"Metric": "NEW", "Value": summary["new"]},
        {"Metric": "RECURRING", "Value": summary["recurring"]},
        {"Metric": "RESOLVED", "Value": summary["resolved"]},
        {"Metric": "DATA", "Value": PROVENANCE_TEXT[data_mode]},
    )


def build_register_workbook(
    result: ReconciliationResult,
    *,
    data_mode: Literal["library", "upload"] = "library",
) -> Workbook:
    """Build the M365-shaped register workbook in memory."""

    workbook = Workbook()
    mismatch_sheet = workbook.active
    mismatch_sheet.title = REGISTER_SHEET
    _append_rows(mismatch_sheet, FINDING_COLUMNS, result.findings)
    _style_tabular_sheet(
        mismatch_sheet,
        text_columns=("Requisition No", "Position Number"),
    )

    summary_sheet = workbook.create_sheet(SUMMARY_SHEET)
    _append_rows(summary_sheet, ("Metric", "Value"), _summary_rows(result, data_mode))
    _style_tabular_sheet(summary_sheet)

    if result.resolved:
        resolved_sheet = workbook.create_sheet(RESOLVED_SHEET)
        _append_rows(resolved_sheet, RESOLVED_COLUMNS, result.resolved)
        _style_tabular_sheet(resolved_sheet, text_columns=("Requisition No",))

    return workbook


def write_register(
    result: ReconciliationResult,
    destination: WorkbookDestination,
    *,
    data_mode: Literal["library", "upload"] = "library",
) -> None:
    """Write a reconciliation register to a path or binary file-like object."""

    workbook = build_register_workbook(result, data_mode=data_mode)
    workbook.save(destination)


def register_to_bytes(
    result: ReconciliationResult,
    *,
    data_mode: Literal["library", "upload"] = "library",
) -> bytes:
    """Return a download-ready register workbook."""

    destination = BytesIO()
    write_register(result, destination, data_mode=data_mode)
    return destination.getvalue()
