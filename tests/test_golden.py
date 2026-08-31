from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from engine.io_excel import (
    PROVENANCE_TEXT,
    REGISTER_SHEET,
    RESOLVED_SHEET,
    SUMMARY_SHEET,
    HeaderValidationError,
    load_inputs,
    read_config,
    read_extract,
    read_previous_findings,
    reconcile_workbooks,
    register_to_bytes,
    validate_headers,
)
from engine.library import discover_library_extracts, run_library as run_library_engine
from engine.reconcile import FINDING_COLUMNS, ReconciliationResult, reconcile


ROOT = Path(__file__).resolve().parents[1]
LIBRARY = ROOT / "synthetic_library"
CONFIG_PATH = LIBRARY / "config" / "config_field_map.xlsx"
DATES = ("2026-08-20", "2026-08-21", "2026-08-24", "2026-08-25", "2026-08-26")

EXPECTED = {
    "2026-08-20": {"active_count": 55, "exceptions": 3, "new": 3, "recurring": 0, "resolved_ids": set()},
    "2026-08-21": {"active_count": 56, "exceptions": 4, "new": 1, "recurring": 3, "resolved_ids": set()},
    "2026-08-24": {
        "active_count": 56,
        "exceptions": 2,
        "new": 0,
        "recurring": 2,
        "resolved_ids": {"101103", "101109"},
    },
    "2026-08-25": {
        "active_count": 57,
        "exceptions": 3,
        "new": 3,
        "recurring": 0,
        "resolved_ids": {"101112", "101131"},
    },
    "2026-08-26": {
        "active_count": 58,
        "exceptions": 3,
        "new": 1,
        "recurring": 2,
        "resolved_ids": {"101118"},
    },
}


def workbook_paths(run_date: str) -> tuple[Path, Path]:
    folder = LIBRARY / "extracts" / run_date
    return (
        folder / f"RCM_Requisition_Report_{run_date}.xlsx",
        folder / f"EC_Position_Report_{run_date}.xlsx",
    )


def run_library(*, apply_normalisers: bool = True):
    return run_library_engine(LIBRARY, apply_normalisers=apply_normalisers)


def finding_for(result, requisition_no: str):
    matches = [item for item in result.findings if item["Requisition No"] == requisition_no]
    assert len(matches) == 1
    return matches[0]


def test_chained_library_matches_golden_table_exactly():
    results = run_library()
    for run_date, expected in EXPECTED.items():
        result = results[run_date]
        assert result.summary["active_count"] == expected["active_count"]
        assert result.summary["exceptions"] == expected["exceptions"]
        assert result.summary["new"] == expected["new"]
        assert result.summary["recurring"] == expected["recurring"]
        assert {item["Requisition No"] for item in result.resolved} == expected["resolved_ids"]
        assert all(tuple(item) == FINDING_COLUMNS for item in result.findings)


def test_library_discovery_is_chronological_and_complete():
    extracts = discover_library_extracts(LIBRARY)
    assert tuple(extract.run_date for extract in extracts) == DATES


def test_golden_story_beats_and_point_assertions():
    results = run_library()

    day_21_new = [item for item in results["2026-08-21"].findings if item["Status"] == "New"]
    assert [(item["Requisition No"], item["Exception Type"]) for item in day_21_new] == [
        ("101109", "ORPHAN REQUISITION (no position row)")
    ]

    day_25_new_ids = {
        item["Requisition No"] for item in results["2026-08-25"].findings if item["Status"] == "New"
    }
    assert day_25_new_ids == {"101118", "101127", "101140"}

    item_101140 = finding_for(results["2026-08-25"], "101140")
    assert item_101140["Field"] == "Hiring Manager Name <> Line Manager Name"
    assert item_101140["Exception Type"] == "MISMATCH"
    assert item_101140["Severity"] == "HIGH"
    assert item_101140["Recruitment Stage"] == "Offer"

    item_101127 = finding_for(results["2026-08-26"], "101127")
    assert item_101127["Field"] == "FBS Function <> Function Name"
    assert item_101127["Exception Type"] == "MISMATCH"
    assert item_101127["Severity"] == "MEDIUM"
    assert item_101127["Recruitment Stage"] == "Screening"
    assert item_101127["Status"] == "Recurring"
    assert "People &amp; Culture Operations ⟷ People Operations" in next(
        digest.html for digest in results["2026-08-26"].digests.values() if "101127" in digest.html
    )

    item_101144 = finding_for(results["2026-08-26"], "101144")
    assert item_101144["Field"] == "Cost Center Number <> Cost Center Code"
    assert item_101144["Requisition Value"] == "45888"
    assert item_101144["Position Value"] == "45008"
    assert item_101144["Severity"] == "HIGH"
    assert item_101144["Recruitment Stage"] == "Offer"
    assert item_101144["Status"] == "New"

    recurring_ids = {
        item["Requisition No"] for item in results["2026-08-26"].findings if item["Status"] == "Recurring"
    }
    assert recurring_ids == {"101127", "101140"}


def test_normaliser_off_surfaces_country_false_positives():
    results = run_library(apply_normalisers=False)
    final = results["2026-08-26"]
    country_findings = [item for item in final.findings if item["Field"] == "Country <> Country"]
    assert len(final.findings) == 61
    assert len(country_findings) == 58


@pytest.mark.parametrize(
    ("report", "missing_column", "message_label"),
    (
        ("requisition", "Cost Center Number", "Requisition report"),
        ("position", "Line Manager Name", "Position report"),
    ),
)
def test_header_validation_names_missing_required_column(report, missing_column, message_label):
    req_path, pos_path = workbook_paths("2026-08-26")
    requisitions = read_extract(req_path)
    positions = read_extract(pos_path)
    config = read_config(CONFIG_PATH)
    if report == "requisition":
        requisitions = requisitions.drop(columns=[missing_column])
    else:
        positions = positions.drop(columns=[missing_column])

    with pytest.raises(HeaderValidationError) as caught:
        validate_headers(requisitions, positions, config)

    assert str(caught.value) == f"{message_label} missing required columns: {missing_column}"


def test_excel_reader_preserves_text_ids_with_leading_zeroes():
    workbook = BytesIO()
    source = pd.DataFrame(
        {
            "Position Number": pd.Series(["001234"], dtype="string"),
            "Cost Center Number": pd.Series(["0045008"], dtype="string"),
        }
    )
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        source.to_excel(writer, sheet_name="Report", index=False)
    workbook.seek(0)
    loaded = read_extract(workbook)
    assert loaded.loc[0, "Position Number"] == "001234"
    assert loaded.loc[0, "Cost Center Number"] == "0045008"
    assert loaded.dtypes.astype(str).to_dict() == {
        "Position Number": "object",
        "Cost Center Number": "object",
    }


def test_engine_module_does_not_import_streamlit():
    import engine.reconcile as engine_module

    assert "streamlit" not in engine_module.__dict__


def _summary_values(workbook):
    sheet = workbook[SUMMARY_SHEET]
    return {sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=2).value for row in range(2, sheet.max_row + 1)}


def test_register_writer_has_exact_sheets_columns_summary_and_text_ids():
    result = run_library()["2026-08-26"]
    register = register_to_bytes(result, data_mode="library")
    workbook = load_workbook(BytesIO(register), data_only=True)

    assert workbook.sheetnames == [REGISTER_SHEET, SUMMARY_SHEET, RESOLVED_SHEET]
    mismatch = workbook[REGISTER_SHEET]
    assert tuple(cell.value for cell in mismatch[1]) == FINDING_COLUMNS
    assert mismatch.max_row == 4
    assert all(mismatch.cell(row=row, column=1).data_type == "s" for row in range(2, 5))
    assert all(mismatch.cell(row=row, column=1).number_format == "@" for row in range(2, 5))
    assert all(mismatch.cell(row=row, column=2).data_type == "s" for row in range(2, 5))
    assert all(mismatch.cell(row=row, column=2).number_format == "@" for row in range(2, 5))

    summary = _summary_values(workbook)
    assert summary == {
        "RUN DATE": "2026-08-26",
        "ACTIVE REQUISITIONS": 58,
        "FIELD PAIRS CHECKED": 13,
        "EXCEPTIONS": 3,
        "HIGH": 2,
        "NEW": 1,
        "RECURRING": 2,
        "RESOLVED": 1,
        "DATA": PROVENANCE_TEXT["library"],
    }

    resolved = workbook[RESOLVED_SHEET]
    assert tuple(cell.value for cell in resolved[1]) == (
        "Requisition No",
        "Field",
        "Exception Type",
        "Status",
    )
    assert resolved["A2"].value == "101118"
    assert resolved["A2"].data_type == "s"
    assert resolved["A2"].number_format == "@"


def test_register_omits_empty_resolved_sheet_and_uses_upload_provenance():
    result = run_library()["2026-08-20"]
    register = register_to_bytes(result, data_mode="upload")
    workbook = load_workbook(BytesIO(register), data_only=True)

    assert workbook.sheetnames == [REGISTER_SHEET, SUMMARY_SHEET]
    assert RESOLVED_SHEET not in workbook.sheetnames
    assert _summary_values(workbook)["DATA"] == PROVENANCE_TEXT["upload"]


def test_register_round_trip_preserves_leading_zero_identifiers():
    original = run_library()["2026-08-26"]
    finding = dict(original.findings[0])
    finding["Requisition No"] = "00101127"
    finding["Position Number"] = "00050027"
    result = ReconciliationResult(
        findings=(finding,),
        resolved=(),
        digests={},
        summary={**original.summary, "exceptions": 1, "new": 0, "recurring": 1, "resolved": 0},
    )

    register = register_to_bytes(result)
    reloaded = read_previous_findings(BytesIO(register))
    assert reloaded[0]["Requisition No"] == "00101127"
    assert reloaded[0]["Position Number"] == "00050027"


def test_written_day_n_register_reproduces_day_n_plus_one_chain_statuses():
    in_memory = run_library()
    day_25_register = register_to_bytes(in_memory["2026-08-25"])
    req_path, pos_path = workbook_paths("2026-08-26")

    from_uploaded_previous = reconcile_workbooks(
        req_path,
        pos_path,
        CONFIG_PATH,
        "2026-08-26",
        BytesIO(day_25_register),
    )

    assert from_uploaded_previous.findings == in_memory["2026-08-26"].findings
    assert from_uploaded_previous.resolved == in_memory["2026-08-26"].resolved
    assert from_uploaded_previous.summary == in_memory["2026-08-26"].summary
